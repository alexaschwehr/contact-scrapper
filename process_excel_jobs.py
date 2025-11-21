import os
import json
import sys
from typing import List, Dict, Any
from dotenv import load_dotenv
import openpyxl
from agent import RecruiterFinderAgent
from salesql_enricher import enrich_executive_profiles

load_dotenv()


def read_job_urls_from_excel(file_path: str = "linkedin_jobpost.xlsx") -> List[str]:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        urls = []

        header_row = 1
        url_col_idx = None
        

        for col_idx, cell in enumerate(ws[header_row], 1):
            cell_value = str(cell.value or "").lower()
            if "url" in cell_value or "link" in cell_value:
                url_col_idx = col_idx
                break

        if url_col_idx is None:
            url_col_idx = 1

        start_row = 2 if url_col_idx == 1 and ws[header_row][0].value else 1
        
        for row_idx in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=url_col_idx)
            url = str(cell.value or "").strip()
            
            # Validate URL
            if url and (url.startswith("http://") or url.startswith("https://")):
                urls.append(url)
        
        wb.close()
        return urls
    
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")


def process_jobs_from_excel(
    excel_file: str = "linkedin_jobpost.xlsx",
    output_file: str = "executive_profiles_results.json",
    headless: bool = True,
    enable_email_enrichment: bool = True
) -> None:

    print("="*60)
    print("EXCEL JOB PROCESSOR - Executive Profile Finder")
    print("="*60)
    
    # Read URLs from Excel
    print(f"\n Reading job URLs from: {excel_file}")
    try:
        urls = read_job_urls_from_excel(excel_file)
        print(f" Found {len(urls)} job URL(s) in Excel file")
    except Exception as e:
        print(f" Error reading Excel file: {str(e)}")
        return
    
    if not urls:
        print("  No valid URLs found in Excel file.")
        return
    
    # Initialize agent
    print("\n Initializing agent...")
    try:
        agent = RecruiterFinderAgent()
        print(" Agent initialized successfully")
    except Exception as e:
        print(f" Error initializing agent: {str(e)}")
        return
    
    # Process each URL
    print(f"\n Processing {len(urls)} job posting(s)...")
    print("="*60)
    
    all_results = []
    
    for idx, url in enumerate(urls, 1):
        print(f"\n[{idx}/{len(urls)}] Processing: {url}")
        print("-" * 60)
        
        try:
            result = agent.find_recruiters_for_job(url)
            all_results.append(result)
            
            # Display quick summary
            if "error" in result:
                print(f" Error: {result.get('error', 'Unknown error')}")
            else:
                company = result.get('company', 'N/A')
                ceo = result.get('ceo')
                cto = result.get('cto')
                cfo = result.get('cfo')
                hr_count = len(result.get('hr', []))
                recruiter_count = len(result.get('recruiters', []))
                
                print(f"   Company: {company}")
                print(f"   CEO: {'Found' if ceo else 'Not found'}")
                print(f"   CTO: {'Found' if cto else 'Not found'}")
                print(f"   CFO: {'Found' if cfo else 'Not found'}")
                print(f"   HR: {hr_count} profile(s)")
                print(f"   Recruiters: {recruiter_count} profile(s)")
        
        except KeyboardInterrupt:
            print("\n\n  Interrupted by user.")
            break
        except Exception as e:
            print(f" Error processing URL: {str(e)}")
            all_results.append({
                "job_url": url,
                "error": str(e)
            })
    
    # Save results
    print("\n" + "="*60)
    print(f" Saving results to: {output_file}")
    
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_results, f, indent=2, ensure_ascii=False)
        print(f" Results saved successfully!")
        print(f"   Total jobs processed: {len(all_results)}")
    except Exception as e:
        print(f" Error saving results: {str(e)}")
    
    print("\n" + "="*60)
    print(" Processing complete!")
    
    # Email enrichment step (if enabled)
    if enable_email_enrichment:
        print("\n" + "="*60)
        print(" EMAIL ENRICHMENT - SalesQL Integration")
        print("="*60)
        
        # Check if enrichment should be enabled (based on API key presence)
        enrichment_enabled = os.getenv("SALESQL_ENABLED", "true").lower() == "true"
        
        if enrichment_enabled:
            try:
                enrich_executive_profiles(
                    input_json_path=output_file,
                    enabled=enrichment_enabled
                )
            except Exception as e:
                print(f"  Email enrichment encountered an error: {str(e)}")
                print("   Continuing without email enrichment...")
        else:
            print("  Email enrichment is disabled (SALESQL_ENABLED=false)")
    else:
        print("\n  Email enrichment skipped (disabled in function call)")


def main():
    """Main entry point."""
    excel_file = "linkedin_jobpost.xlsx"
    output_file = "executive_profiles_results.json"
    enable_enrichment = True
    
    # Allow custom file paths via command line
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    if len(sys.argv) > 3:
        # Third argument can disable enrichment: python process_excel_jobs.py file.xlsx output.json false
        enable_enrichment = sys.argv[3].lower() != "false"
    
    # Check environment variable for enrichment flag
    env_enrichment = os.getenv("SALESQL_ENABLED", "true").lower() == "true"
    enable_enrichment = enable_enrichment and env_enrichment
    
    try:
        process_jobs_from_excel(excel_file, output_file, enable_email_enrichment=enable_enrichment)
    except KeyboardInterrupt:
        print("\n\n  Interrupted by user.")
    except Exception as e:
        print(f"\n Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

